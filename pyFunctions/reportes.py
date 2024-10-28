import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from models.Database import getDatabase
from models.Usuario import Usuario
from models.Transaccion import Transaccion
from models.DetalleTransaccion import DetalleTransaccion
from models.Producto import Producto

db = getDatabase()

def generar_informe_ventas_mensual(empleado_id, year, month):
    try:
        # Obtener los datos del empleado
        empleado = Usuario.query.filter_by(id=empleado_id).first()
        if not empleado:
            print("Empleado no encontrado")
            return None

        # Filtrar transacciones del empleado para el mes y año dados
        transacciones = Transaccion.query.filter(
            Transaccion.empleado_id == empleado_id,
            db.extract('year', Transaccion.fecha) == year,
            db.extract('month', Transaccion.fecha) == month
        ).all()

        if not transacciones:
            print("No hay transacciones para este periodo.")
            return None

        # Crear lista para almacenar detalles del reporte
        detalles = []
        monto_total = 0

        # Recorrer las transacciones y sus detalles
        for transaccion in transacciones:
            detalles_transaccion = DetalleTransaccion.query.filter_by(transaccion_id=transaccion.id).all()
            for detalle in detalles_transaccion:
                producto = Producto.query.filter_by(id=detalle.producto_id).first()
                detalles.append({
                    'Fecha de Venta': transaccion.fecha,
                    'Producto': producto.nombre if producto else 'Producto no encontrado',
                    'Cantidad': detalle.cantidad,
                    'Monto Pagado': transaccion.monto
                })
                monto_total += transaccion.monto

        # Convertir detalles a DataFrame
        df_detalles = pd.DataFrame(detalles)

        # Crear resumen como DataFrame
        resumen = pd.DataFrame({
            'Información del Reporte': [
                'Nombre del Empleado:', 
                f"{empleado.nombre} {empleado.apellido}",
                'Fecha del Reporte:', 
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'Mes de Reporte:', 
                f"{month}-{year}",
                'Monto Total de Ventas:', 
                f"${monto_total:.2f}"
            ]
        })

        # Guardar el archivo de reporte temporalmente
        nombre_archivo = f"Reporte_Ventas_{empleado_id}_{year}_{month}.xlsx"
        with pd.ExcelWriter(nombre_archivo, engine='openpyxl') as writer:
            # Agregar un título al documento
            title = "Informe Mensual de Ventas"
            ws = writer.book.create_sheet(title="Ventas")  # Crear nueva hoja
            ws.append([title])
            ws['A1'].font = Font(size=16, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
            ws.merge_cells('A1:B1')  # Fusionar celdas para el título
            
            # Insertar el resumen
            for index in range(0, len(resumen), 2):  # Procesar de dos en dos
                # Encabezado en columna B y valor en columna C
                ws.cell(row=index + 2, column=2, value=resumen.iloc[index, 0])  # Encabezado en B
                ws.cell(row=index + 2, column=3, value=resumen.iloc[index + 1, 0])  # Valor en C

            # Agregar detalles
            ws.append([])  # Añadir una fila en blanco
            df_detalles.to_excel(writer, index=False, startrow=len(resumen) + 3, sheet_name='Ventas')  # Deja espacio para el resumen

        # Cargar el archivo para aplicar estilo
        wb = load_workbook(nombre_archivo)
        ws = wb['Ventas']  # Seleccionar la hoja creada

        # Aplicar estilos al resumen
        for row in range(2, len(resumen) + 2):  # Empezar desde la segunda fila del resumen
            cell = ws.cell(row=row, column=2)
            cell.font = Font(bold=True, size=12)
            cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
            cell.alignment = Alignment(horizontal="left")

        # Ajustar los anchos de columna para la tabla de detalles
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15

        # Aplicar estilo a los encabezados de la tabla de detalles
        for cell in ws[len(resumen) + 4]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Alinear y aplicar formato de moneda a los montos pagados
        for row in range(len(resumen) + 5, ws.max_row + 1):
            ws.cell(row=row, column=4).alignment = Alignment(horizontal="right")
            ws.cell(row=row, column=4).number_format = '"$"#,##0.00'

        # Guardar el archivo con los estilos aplicados
        wb.save(nombre_archivo)

        print(f"Informe generado exitosamente: {nombre_archivo}")
        return nombre_archivo
    except Exception as e:
        print(f"Error al generar el informe: {e}")
        return None
